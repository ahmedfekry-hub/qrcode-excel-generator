import io
import os
import zipfile
import tempfile
import streamlit as st
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont, ImageChops
import qrcode

# =========================
# SETTINGS
# =========================
LABEL_COLS = 3
LABEL_ROWS = 7

QR_SIZE = 260
LOGO_TARGET_W = 260

EXCEL_IMG_W = 260
EXCEL_IMG_H = 460

LABEL_COL_WIDTH = 23
LABEL_ROW_HEIGHT = 360

APP_DIR = os.path.dirname(os.path.abspath(__file__))
LOGO_PATH = os.path.join(APP_DIR, "logo.png")


# =========================
# TRIM LOGO WHITESPACE
# =========================
def trim_logo_whitespace(im):

    if im.mode != "RGBA":
        im = im.convert("RGBA")

    white_bg = Image.new("RGBA", im.size, (255, 255, 255, 255))
    merged = Image.alpha_composite(white_bg, im).convert("RGB")

    diff = ImageChops.difference(merged, Image.new("RGB", merged.size, (255, 255, 255)))
    bbox = diff.getbbox()

    if bbox:
        pad = 6
        x0, y0, x1, y1 = bbox
        x0 = max(0, x0 - pad)
        y0 = max(0, y0 - pad)
        x1 = min(im.size[0], x1 + pad)
        y1 = min(im.size[1], y1 + pad)
        return im.crop((x0, y0, x1, y1))

    return im


# =========================
# CREATE QR + LOGO BLOCK
# =========================
def make_qr_block(qr_data, building_id, out_png):

    logo = Image.open(LOGO_PATH).convert("RGBA")
    logo = trim_logo_whitespace(logo)

    # Force logo width = QR width
    scale = LOGO_TARGET_W / logo.width
    logo = logo.resize((LOGO_TARGET_W, int(logo.height * scale)), Image.LANCZOS)

    qr = qrcode.QRCode(
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=2,
    )

    qr.add_data(qr_data)
    qr.make(fit=True)

    qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGBA")
    qr_img = qr_img.resize((QR_SIZE, QR_SIZE), Image.NEAREST)

    canvas_w = 320
    canvas_h = 520

    canvas = Image.new("RGBA", (canvas_w, canvas_h), "white")

    # paste logo
    canvas.paste(logo, ((canvas_w - logo.width) // 2, 10), logo)

    qr_y = 10 + logo.height + 12

    # paste qr
    canvas.paste(qr_img, ((canvas_w - QR_SIZE) // 2, qr_y), qr_img)

    draw = ImageDraw.Draw(canvas)

    orange = (255, 140, 0)

    font = None

    for fp in [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "C:/Windows/Fonts/arialbd.ttf",
        "C:/Windows/Fonts/calibrib.ttf",
    ]:
        try:
            font = ImageFont.truetype(fp, 28)
            break
        except:
            pass

    if font is None:
        font = ImageFont.load_default()

    text = str(building_id).strip()

    bbox = draw.textbbox((0, 0), text, font=font)
    tw = bbox[2] - bbox[0]
    th = bbox[3] - bbox[1]

    draw.text(
        ((canvas_w - tw) // 2, canvas_h - th - 18),
        text,
        fill=orange,
        font=font,
    )

    canvas.convert("RGB").save(out_png, "PNG")


# =========================
# FIND HEADER ROW
# =========================
def find_header_row(ws):

    for r in range(1, 60):

        row_vals = [
            str(ws.cell(r, c).value or "").strip().lower()
            for c in range(1, ws.max_column + 1)
        ]

        if "building code" in row_vals or "building id" in row_vals:
            return r

    return None


# =========================
# DETECT COLUMNS
# =========================
def detect_columns(ws, header_row):

    headers = {}

    for c in range(1, ws.max_column + 1):

        v = ws.cell(header_row, c).value

        if v:
            headers[str(v).strip().lower()] = c

    def exact(names):
        for n in names:
            if n in headers:
                return headers[n]
        return None

    def contains(keys):
        for k, v in headers.items():
            for key in keys:
                if key in k:
                    return v
        return None

    col_building = exact(["building code", "building id"]) or contains(
        ["building code", "building id"]
    )

    col_address = exact(["national address"]) or contains(
        ["national address", "address"]
    )

    col_qr = exact(["barcode", "qr"]) or contains(["barcode", "qr"])

    if col_building is None:
        return None

    if col_address is None:
        col_address = min(col_building + 1, ws.max_column)

    if col_qr is None:
        col_qr = ws.max_column + 1
        ws.cell(header_row, col_qr, "Barcode")

    return col_building, col_address, col_qr


# =========================
# SETUP LABELS SHEET
# =========================
def setup_labels_sheet(wb):

    if "LABELS" in wb.sheetnames:
        del wb["LABELS"]

    ws = wb.create_sheet("LABELS")

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

    for c in range(1, LABEL_COLS + 1):
        ws.column_dimensions[get_column_letter(c)].width = LABEL_COL_WIDTH

    return ws


# =========================
# BUILDING ID CHECK
# =========================
def looks_like_building_id(x):

    s = str(x).strip()

    if len(s) < 6:
        return False

    if s[0].isalpha() and len(s) >= 8:
        return True

    if s.isdigit() and len(s) >= 8:
        return True

    return False


# =========================
# PROCESS FILE
# =========================
def process_xlsx(xlsx_bytes, filename):

    with tempfile.TemporaryDirectory() as td:

        src = os.path.join(td, filename)

        with open(src, "wb") as f:
            f.write(xlsx_bytes)

        wb = load_workbook(src)

        labels_ws = setup_labels_sheet(wb)

        all_images = []

        for ws in wb.worksheets:

            if ws.title.strip().upper() == "LABELS":
                continue

            header_row = find_header_row(ws)

            if header_row is None:
                continue

            cols = detect_columns(ws, header_row)

            if cols is None:
                continue

            col_building, col_address, col_qr = cols

            try:
                ws._images = []
            except:
                pass

            qr_col_letter = get_column_letter(col_qr)

            ws.column_dimensions[qr_col_letter].width = 28

            for r in range(header_row + 1, ws.max_row + 1):

                bid = ws.cell(r, col_building).value

                if not bid or not looks_like_building_id(bid):
                    continue

                addr = ws.cell(r, col_address).value

                qr_data = str(bid).strip()

                if addr:
                    qr_data += "\n" + str(addr).strip()

                img_path = os.path.join(td, f"{ws.title}_{r}.png")

                make_qr_block(qr_data, bid, img_path)

                all_images.append(img_path)

                img = XLImage(img_path)

                img.width = EXCEL_IMG_W
                img.height = EXCEL_IMG_H

                ws.add_image(img, f"{qr_col_letter}{r}")

                ws.row_dimensions[r].height = 360

        per_page = LABEL_COLS * LABEL_ROWS

        for i, img_path in enumerate(all_images):

            page = i // per_page
            pos = i % per_page

            row = page * (LABEL_ROWS + 1) + (pos // LABEL_COLS) + 1
            col = (pos % LABEL_COLS) + 1

            labels_ws.row_dimensions[row].height = LABEL_ROW_HEIGHT

            cell = f"{get_column_letter(col)}{row}"

            img = XLImage(img_path)

            img.width = EXCEL_IMG_W
            img.height = EXCEL_IMG_H

            labels_ws.add_image(img, cell)

        out = io.BytesIO()

        wb.save(out)

        return out.getvalue()


# =========================
# STREAMLIT UI
# =========================
st.set_page_config(page_title="QR Excel Generator", layout="centered")

st.title("QR Code Excel Generator (Logo + A4 Labels)")

files = st.file_uploader(
    "Upload Excel files (.xlsx)",
    type=["xlsx"],
    accept_multiple_files=True,
)

if files:

    if st.button("Generate"):

        zip_buf = io.BytesIO()

        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as z:

            for f in files:

                result = process_xlsx(f.read(), f.name)

                out_name = f.name.replace(".xlsx", "_QR_READY.xlsx")

                z.writestr(out_name, result)

        st.success("Done ✅")

        st.download_button(
            "Download ZIP (QR_READY files)",
            zip_buf.getvalue(),
            file_name="QR_READY_OUTPUT.zip",
            mime="application/zip",
        )

else:
    st.info("Upload one or more Excel files to begin.")
